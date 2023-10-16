from robocorp.tasks import task
from RPA.Excel.Files import Files
from RPA.Excel.Application import Application
from openpyxl import load_workbook
import os
import xml.etree.ElementTree as ET
import sys
from lxml import etree as ET
import shutil
from datetime import datetime, timedelta
import fnmatch

@task
def main():
    check_files_and_execute_code()  #checks if the folder runs_totP has results named _totP


def check_files_and_execute_code():
    directory = "C:\\Tytti\\Fosfori\\ajot_totP"
    destination = "C:\\Tytti\\Fosfori\\ajot_totP\\xml_luotu"
    pattern = "*_totP*"

    # Initialize the found variable to False.
    found = False

    # Loop through the files in the directory
    for file_name in os.listdir(directory):
            # If a match is found, set found to True and execute other functions.
            if fnmatch.fnmatch(file_name, pattern):
                found = True
                full_file_path = os.path.join(directory, file_name)
                copy_rows_from_ajot_to_pohja(full_file_path)    # Open ajot.xlsx and read the data from 'kokP' and paste it to pohja_totP. Before paste clear the sheet1.
                shutil.move(full_file_path, destination)  # Move the file to the destination folder
                run_macro() # Use macros from macro.xlsm and produce totP.xml
                compare_results()   # Compare totP results to earlier determined results in syotetyt folder and inform if totP is lower than before determined result PO4
                rename_xml_and_move_it()    # Rename totP.xml with jobs inside it and remove it to syotetyt folder
                archieving()    # Archieve more than one month older files from syotetyt folder to arkistoitu folder. 

    # If no match was found, write the message to the info_to_user.txt file.
    if not found:
        with open('C:/Tytti/Fosfori/info_to_user.txt', 'a') as f:
            now = datetime.now()
            timestamp = now.strftime("%Y-%m-%d %H:%M:%S")
            f.write(f'{timestamp}: Ei siirrettäviä totP ajoja\n')

def copy_rows_from_ajot_to_pohja(file_name):
    excel = Files()
    try: 
        # Open ajot.xlsx and read the data from 'kokP' 
        excel.open_workbook(file_name)
        excel.set_cell_value(row=1, column="A", value='A')
        excel.set_cell_value(row=1, column="B", value='B')
        excel.set_cell_value(row=1, column="C", value='C')
        excel.set_cell_value(row=1, column="D", value='D')
        excel.set_cell_value(row=1, column="E", value='E')
        excel.set_cell_value(row=1, column="F", value='F')
        data = excel.read_worksheet_as_table(name='kokP', header=True)
    except Exception as e:
        with open('output.txt', 'a') as f:
            f.write(f"An error occurred while opening ajot.xlsx and reading the data from 'kokP': {e}\n")
        sys.exit(1)
    finally:
        excel.close_workbook()

    # Filter the data to only include columns A-F
    filtered_data = [{k: v for k, v in row.items() if k in ['A', 'B', 'C', 'D', 'E', 'F']} for row in data]

    # Open the pohja_totP workbook using openpyxl, clear the data from sheet1
    wb = load_workbook('C:\\Tytti\\Fosfori\\pohja_totP.xlsx')
    ws = wb['Sheet1']
    for row in ws['A1:F' + str(ws.max_row)]:
        for cell in row:
            cell.value = None

    # Save the workbook
    wb.save('C:\\Tytti\\Fosfori\\pohja_totP.xlsx')

    # Open the pohja_totP workbook,  write the data to sheet1
    try:
        excel.open_workbook('C:\\Tytti\\Fosfori\\pohja_totP.xlsx')
        excel.append_rows_to_worksheet(filtered_data, name='Sheet1', header=True)
        excel.save_workbook()
        excel.close_workbook()
    except Exception as e:
        with open('output.txt', 'a') as f:
            f.write(f"An error occurred while pohja_totP and write data to the sheet1: {e}\n")
        sys.exit(1)
    finally:
        excel.close_workbook()

def run_macro():
    try:
        excel_app = Application()  # Instantiate the Application class
        # Use RPA.Excel.Application to run the macro
        excel_app.open_workbook('C:\\Tytti\Fosfori\\macrot.xlsm')
        excel_app.run_macro('PoistaRivitJaLuoXML')
        excel_app.run_macro('RemoveTabulatorSpacesInXML')
        excel_app.save_excel()
        excel_app.quit_application()
    except Exception as e:
        with open('output.txt', 'a') as f:
            f.write(f"An error occurred running macros: {e}\n")
        sys.exit(1)
    finally:
        excel_app.quit_application()
    
def compare_results():
    # Parse the XML file
    tree = ET.parse('C:\\Tytti\\Fosfori\\totP.xml')
    root = tree.getroot()

    # Find 'NA' elements
    NA_elements = root.findall('.//NA')

    # Open the output file
    with open('C:\Tytti\Fosfori\siirretyt.txt', 'w') as f:
        # Print 'NA' elements
        for NA_element in NA_elements:
                f.write(f"NA element: {NA_element.text}\n")

        # Get the names of all files in the folder
        folder_files = os.listdir('C:\\Tytti\\Fosfori\\syotetyt')

        # Print file names
        with open('C:/Tytti/Fosfori/eteneminen.txt', 'w') as et:
            for file in folder_files:
                et.write(f"File name: {file}\n")

    # Open the output file in write mode
        with open('C:\\Tytti\Fosfori\\totP_vs_PO4.txt', 'w') as f:
            # Check if 'syotetyt' directory exists
            if not os.path.exists('C:\\Tytti\\Fosfori\\syotetyt'):
                    with open('C:/Tytti/Fosfori/eteneminen.txt', 'a') as et:                   
                        et.write("'C:\\Tytti\\Fosfori\\syotetyt' directory does not exist.")
            else:
                with open('C:/Tytti/Fosfori/eteneminen.txt', 'a') as et:
                    et.write("'C:\\Tytti\\Fosfori\\syotetyt' directory exists.")
                # Open the totP.xml file and find the 'NA' elements
                with open('C:\\Tytti\\Fosfori\\totP.xml', 'r') as totP_file:
                    tree = ET.parse(totP_file)
                    root = tree.getroot()
                    for NA_element in root.iter('NA'):
                        with open('C:/Tytti/Fosfori/eteneminen.txt', 'a') as et:
                            et.write(f"Processing NA element: {NA_element.text}\n")
                        # Iterate over the files in the 'ajetut' folder
                        for file in os.listdir('C:\\Tytti\\Fosfori\\syotetyt'):
                            with open('C:/Tytti/Fosfori/eteneminen.txt', 'a') as et:
                                et.write(f"Checking file: {file}\n")
                            # Split the NA element text on '/' and take the first part
                            NA_element_part = NA_element.text.split('/')[0]

                            # Check if the NA element part is in the file name
                            if NA_element_part in file:
                                with open('C:/Tytti/Fosfori/eteneminen.txt', 'a') as et:
                                    et.write(f"Found matching file: {file}\n")  
                                with open(f'C:\\Tytti\\Fosfori\\syotetyt/{file}', 'r') as specific_file:
                                    specific_tree = ET.parse(specific_file)
                                    specific_root = specific_tree.getroot()
                                    for specific_NA_element in specific_root.iter('NA'):
                                        if specific_NA_element.text == NA_element.text:
                                            totP_result_element = NA_element.getparent().getparent().find('DATA/Result')
                                            specific_result_element = specific_NA_element.getparent().getparent().find('DATA/Result')
                                            if totP_result_element is not None and specific_result_element is not None:
                                                totP_result = float(totP_result_element.text)
                                                specific_result = float(specific_result_element.text)
                                                if totP_result < specific_result:
                                                    with open('C:\\Tytti\Fosfori\\totP_vs_PO4.txt', 'a') as f:
                                                        f.write(f"\n The Result {totP_result} of {NA_element.text} in {file} is lower than the result {specific_result} in totP.xml \n")
                                            else:
                                                print("Couldn't find 'DATA/Result' in SAMPLE")

def rename_xml_and_move_it():
    # Parse the XML file
    tree = ET.parse('C:\\Tytti\\Fosfori\\totP.xml')
    root = tree.getroot()

    # Find all 'NA' elements
    na_elements = root.findall('.//NA')

    # Get the names of all 'NA' elements, only including characters before '/'
    na_names = [na.text.split('/')[0] for na in na_elements]

    # Remove duplicates by converting the list to a set and back to a list
    na_names = list(set(na_names))

    # Specify the full path of the file
    file_path = 'C:\\Tytti\\Fosfori\\totP.xml'

    # Create the new file name with the full path
    new_file_name = 'C:\\Tytti\\Fosfori\\totP_' + '_'.join(na_names) + '.xml'

    # Rename the file
    os.rename(file_path, new_file_name)

    # Get the base name of the new file
    new_file_base_name = os.path.basename(new_file_name)

    # Move the file to the new directory
    shutil.move(new_file_name, 'C:\\Tytti\\Fosfori\\syotetyt\\' + new_file_base_name)

def archieving():
    # Define the source and destination directories
    source_dir = 'C:\\Tytti\\Fosfori\\syotetyt'
    dest_dir = 'C:\\Tytti\\Fosfori\\arkistoitu'

    # Get the current date and time
    now = datetime.now()

    # Define the cutoff for old files (1 month ago)
    cutoff = now - timedelta(days=30)

    # Iterate over the files in the source directory
    for filename in os.listdir(source_dir):
        # Get the full path of the file
        file_path = os.path.join(source_dir, filename)
    
        # Get the last modified date of the file
        modified_date = datetime.fromtimestamp(os.path.getmtime(file_path))
    
        # If the file was modified more than a month ago, move it to the destination directory
        if modified_date < cutoff:
            shutil.move(file_path, dest_dir)   


if __name__ == "__main__":
    main()